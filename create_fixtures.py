"""Script to generate test fixture files."""
from __future__ import annotations

import sys
from pathlib import Path

# Ensure we can import openpyxl
try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    print("Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

FIXTURES_DIR = Path(__file__).parent / "tests" / "fixtures"
FIXTURES_DIR.mkdir(parents=True, exist_ok=True)


def create_sample_rfp_xlsx() -> None:
    """Create a minimal fixture that mirrors the real RFP structure."""
    wb = openpyxl.Workbook()

    # ---- Tracker sheet ----
    tracker = wb.active
    tracker.title = "Tracker"
    tracker["A1"] = "Section"
    tracker["B1"] = "Count Answered"
    tracker["A2"] = "Business Architecture"
    # COUNTA formula referencing Technical (App)!F5:F7
    tracker["B2"] = "=COUNTA('Technical (App)'!F5:F7)"
    tracker["A3"] = "Data Architecture"
    tracker["B3"] = "=COUNTA('Technical (App)'!F9:F10)"

    # ---- Technical (App) sheet ----
    tech = wb.create_sheet("Technical (App)")

    # Row 1: empty (as per real file)
    # Row 2: headers
    headers = ["", "No", "Label", "Type", "Response", "Comments"]
    for col, val in enumerate(headers, start=1):
        tech.cell(row=2, column=col).value = val

    # Row 3-4: empty spacers

    # Row 5: section header (B empty, D=Comment, C=section name)
    tech.cell(row=5, column=2).value = None           # B empty
    tech.cell(row=5, column=3).value = "Business Architecture"
    tech.cell(row=5, column=4).value = "Comment"
    # B is empty → section header

    # Rows 5-7: questions
    questions = [
        ("7.1", "Does the solution support multi-tenancy?", "Single Choice"),
        ("7.2", "Describe the deployment model.", "Comment"),
        ("7.3", "Is the solution cloud-native?", "Single Choice"),
    ]
    for i, (no, label, qtype) in enumerate(questions):
        row = 6 + i   # rows 6,7,8
        tech.cell(row=row, column=2).value = no
        tech.cell(row=row, column=3).value = label
        tech.cell(row=row, column=4).value = qtype

    # Row 9: blank spacer
    # Row 9: another section header
    tech.cell(row=9, column=2).value = None
    tech.cell(row=9, column=3).value = "Data Architecture"
    tech.cell(row=9, column=4).value = "Comment"

    # Rows 10-11: more questions
    questions2 = [
        ("8.1", "How is data encrypted at rest?", "Single Choice"),
        ("8.2", "Describe data residency controls.", "Comment"),
    ]
    for i, (no, label, qtype) in enumerate(questions2):
        row = 10 + i
        tech.cell(row=row, column=2).value = no
        tech.cell(row=row, column=3).value = label
        tech.cell(row=row, column=4).value = qtype

    out = FIXTURES_DIR / "sample_rfp.xlsx"
    wb.save(str(out))
    print(f"Created: {out}")


def create_sample_txt() -> None:
    txt = FIXTURES_DIR / "sample.txt"
    txt.write_text(
        "BusinessNext supports multi-tenancy through logical data isolation per tenant. "
        "Each tenant's data is stored in a separate schema within the shared database. "
        "The platform is deployed on AWS and supports private cloud (on-premise) as well. "
        "Data at rest is encrypted using AES-256. Data in transit uses TLS 1.2+.",
        encoding="utf-8",
    )
    print(f"Created: {txt}")


if __name__ == "__main__":
    create_sample_rfp_xlsx()
    create_sample_txt()
    print("All fixtures created.")
