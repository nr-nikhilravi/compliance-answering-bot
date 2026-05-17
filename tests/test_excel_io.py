"""Tests for Excel I/O — read/write, Tracker formula preservation, row classification."""
from __future__ import annotations

import pytest
from pathlib import Path

import openpyxl

from rfp_responder.config import AppConfig
from rfp_responder.excel_io import read_rfp_sheet, write_output_excel, ROW_TYPE_QUESTION, ROW_TYPE_SECTION

FIXTURES = Path(__file__).parent / "fixtures"


@pytest.fixture
def sample_rfp_path():
    p = FIXTURES / "sample_rfp.xlsx"
    if not p.exists():
        pytest.skip("Fixture not found — run create_fixtures.py first")
    return p


class TestReadRfpSheet:
    def test_reads_questions(self, sample_rfp_path):
        wb = openpyxl.load_workbook(str(sample_rfp_path))
        cfg = AppConfig()
        rows = read_rfp_sheet(wb, cfg)
        questions = [r for r in rows if r.row_type == ROW_TYPE_QUESTION]
        assert len(questions) == 5  # 7.1, 7.2, 7.3, 8.1, 8.2

    def test_detects_section_headers(self, sample_rfp_path):
        wb = openpyxl.load_workbook(str(sample_rfp_path))
        cfg = AppConfig()
        rows = read_rfp_sheet(wb, cfg)
        sections = [r for r in rows if r.row_type == ROW_TYPE_SECTION]
        assert len(sections) >= 2
        section_names = [r.question_text for r in sections]
        assert "Business Architecture" in section_names
        assert "Data Architecture" in section_names

    def test_section_headers_not_in_questions(self, sample_rfp_path):
        wb = openpyxl.load_workbook(str(sample_rfp_path))
        cfg = AppConfig()
        rows = read_rfp_sheet(wb, cfg)
        questions = [r for r in rows if r.row_type == ROW_TYPE_QUESTION]
        labels = [r.question_text for r in questions]
        # Section headers should not appear as question text
        assert "Business Architecture" not in labels

    def test_question_numbers_parsed(self, sample_rfp_path):
        wb = openpyxl.load_workbook(str(sample_rfp_path))
        cfg = AppConfig()
        rows = read_rfp_sheet(wb, cfg)
        questions = [r for r in rows if r.row_type == ROW_TYPE_QUESTION]
        q_numbers = [r.question_number for r in questions]
        assert "7.1" in q_numbers
        assert "8.2" in q_numbers


class TestWriteOutputExcel:
    def test_tracker_formulas_preserved(self, sample_rfp_path, tmp_path):
        """The Tracker sheet formula strings must survive round-trip."""
        wb_input = openpyxl.load_workbook(str(sample_rfp_path))
        cfg = AppConfig()
        rows = read_rfp_sheet(wb_input, cfg)
        wb_input.close()

        results = {}
        for r in rows:
            if r.row_type == ROW_TYPE_QUESTION:
                results[r.row_number] = {
                    "single_choice_value": "Yes",
                    "answer_text": "Test answer.",
                    "confidence": "high",
                    "top_retrieval_score": 0.85,
                    "sources": ["doc.pdf"],
                    "needs_review": False,
                    "review_notes": "",
                    "revision_count": 0,
                }

        cfg.paths.output_excel_folder = tmp_path
        out = tmp_path / "out.xlsx"
        write_output_excel(sample_rfp_path, out, results, cfg)

        assert out.exists()

        # Check Tracker formula preserved
        wb_out = openpyxl.load_workbook(str(out))  # NOT data_only
        tracker = wb_out["Tracker"]
        # B2 should still be a COUNTA formula string
        b2_value = tracker["B2"].value
        assert b2_value is not None
        assert "COUNTA" in str(b2_value), f"Expected COUNTA formula in B2, got: {b2_value!r}"

    def test_answers_written_to_column_f(self, sample_rfp_path, tmp_path):
        wb_input = openpyxl.load_workbook(str(sample_rfp_path))
        cfg = AppConfig()
        rows = read_rfp_sheet(wb_input, cfg)
        wb_input.close()

        results = {}
        for r in rows:
            if r.row_type == ROW_TYPE_QUESTION:
                results[r.row_number] = {
                    "single_choice_value": "Yes" if r.question_type == "Single Choice" else "",
                    "answer_text": f"Answer for {r.question_number}",
                    "confidence": "medium",
                    "top_retrieval_score": 0.5,
                    "sources": [],
                    "needs_review": False,
                    "review_notes": "",
                    "revision_count": 0,
                }

        cfg.paths.output_excel_folder = tmp_path
        out = tmp_path / "out_f.xlsx"
        write_output_excel(sample_rfp_path, out, results, cfg)

        wb_out = openpyxl.load_workbook(str(out), data_only=True)
        tech = wb_out["Technical (App)"]

        # Spot-check one question row — row 6 = question 7.1
        f6 = tech.cell(row=6, column=6).value
        assert f6 == "Answer for 7.1"

    def test_audit_columns_written(self, sample_rfp_path, tmp_path):
        wb_input = openpyxl.load_workbook(str(sample_rfp_path))
        cfg = AppConfig()
        rows = read_rfp_sheet(wb_input, cfg)
        wb_input.close()

        results = {}
        for r in rows:
            if r.row_type == ROW_TYPE_QUESTION:
                results[r.row_number] = {
                    "single_choice_value": "",
                    "answer_text": "Audit test",
                    "confidence": "low",
                    "top_retrieval_score": 0.1,
                    "sources": ["a.pdf"],
                    "needs_review": True,
                    "review_notes": "Check this",
                    "revision_count": 1,
                }

        cfg.paths.output_excel_folder = tmp_path
        out = tmp_path / "out_audit.xlsx"
        write_output_excel(sample_rfp_path, out, results, cfg)

        wb_out = openpyxl.load_workbook(str(out), data_only=True)
        tech = wb_out["Technical (App)"]

        # Col G = Confidence (col 7), col J = Needs Review (col 10)
        g6 = tech.cell(row=6, column=7).value
        j6 = tech.cell(row=6, column=10).value
        l6 = tech.cell(row=6, column=12).value

        assert g6 == "low"
        assert j6 == "YES"
        assert l6 == 1
