from __future__ import annotations

"""Excel I/O: read RFP questions and write answered output with audit columns."""

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import column_index_from_string

from .config import AppConfig

logger = logging.getLogger(__name__)

# Audit column headers (appended after col F = Comments)
_AUDIT_HEADERS = ["Confidence", "Top Retrieval Score", "Sources", "Needs Review", "Review Notes", "Revision Count"]

# Color fills and borders
_FILL_GREEN  = PatternFill(fill_type="solid", fgColor="D9F2D9")
_FILL_YELLOW = PatternFill(fill_type="solid", fgColor="FFF2CC")
_FILL_RED    = PatternFill(fill_type="solid", fgColor="FFCCCC")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="000080") # Navy Blue
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_THIN_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

ROW_TYPE_QUESTION = "question"
ROW_TYPE_SECTION  = "section"
ROW_TYPE_EMPTY    = "empty"


@dataclass
class RfpRow:
    """Represents one row from the Technical (App) sheet."""
    row_number: int
    row_type: str          # "question" | "section" | "empty"
    question_number: str   # col B
    question_text: str     # col C
    question_type: str     # col D (Single Choice / Comment)
    existing_response: str # col D
    existing_comment: str  # col E
    generated_comment: str = "" # col F
    comparison_output: str = "" # col G


def read_rfp_sheet(wb: openpyxl.Workbook, cfg: AppConfig) -> list[RfpRow]:
    """
    Read rows from the Technical (App) sheet, classifying each row.
    Returns only section-header and question rows (not empty rows).
    """
    sheet_name = cfg.excel.question_sheet_name
    if sheet_name not in wb.sheetnames:
        # Find the first sheet that is not preserved
        available_sheets = [s for s in wb.sheetnames if s not in cfg.excel.preserve_sheets]
        if available_sheets:
            sheet_name = available_sheets[0]
            logger.info(f"Default sheet not found. Falling back to sheet: {sheet_name}")
        else:
            raise ValueError(f"No valid sheets found in workbook. Available: {wb.sheetnames}")

    ws = wb[sheet_name]

    # Locate column indices from header row (row 2)
    header_row_idx = cfg.excel.header_row
    col_no       = _find_col(ws, header_row_idx, cfg.excel.no_column, ["No", "no", "NO", "Number", "number"])
    col_label    = _find_col(ws, header_row_idx, cfg.excel.question_column, ["Label", "label", "LABEL"])
    col_type     = _find_col(ws, header_row_idx, cfg.excel.type_column, ["Type", "type", "TYPE"])
    col_response = _find_col(ws, header_row_idx, cfg.excel.response_column, ["Response", "response", "RESPONSE"])
    col_existing_comments = _find_col(ws, header_row_idx, cfg.excel.existing_comments_column, ["Comments", "comments", "COMMENTS"])
    col_generated_comments = _find_col(ws, header_row_idx, cfg.excel.generated_comments_column, ["Comments from CB", "comments from cb"])
    col_comparison = _find_col(ws, header_row_idx, cfg.excel.comparison_column, ["Comparison and output", "comparison and output"])

    rows: list[RfpRow] = []
    consecutive_empty = 0
    data_start = header_row_idx + 1
    max_row = ws.max_row or data_start

    for row_idx in range(data_start, max_row + 1):
        no_val    = _cell_str(ws, row_idx, col_no)
        label_val = _cell_str(ws, row_idx, col_label)
        type_val  = _cell_str(ws, row_idx, col_type)
        resp_val  = _cell_str(ws, row_idx, col_response)
        comm_val  = _cell_str(ws, row_idx, col_existing_comments)

        # Stop condition: 3 consecutive empty rows
        if not no_val and not label_val:
            consecutive_empty += 1
            if consecutive_empty >= 3:
                logger.debug("Stopping at row %d — 3 consecutive empty rows", row_idx)
                break
            continue
        else:
            consecutive_empty = 0

        # Classify
        if not no_val and type_val.lower() == "comment":
            row_type = ROW_TYPE_SECTION
        elif no_val:
            row_type = ROW_TYPE_QUESTION
        else:
            row_type = ROW_TYPE_EMPTY

        if row_type == ROW_TYPE_EMPTY:
            continue

        rows.append(RfpRow(
            row_number=row_idx,
            row_type=row_type,
            question_number=no_val,
            question_text=label_val,
            question_type=type_val,
            existing_response=resp_val,
            existing_comment=comm_val,
        ))

    logger.info("Read %d rows (%d questions) from sheet '%s'",
                len(rows),
                sum(1 for r in rows if r.row_type == ROW_TYPE_QUESTION),
                sheet_name)
    return rows


def write_output_excel(
    input_path: Path,
    output_path: Path,
    results: dict[int, dict],  # row_number → field dict
    cfg: AppConfig,
) -> None:
    """
    Open input workbook (preserving all sheets/formulas),
    write answers + audit columns to Technical (App) sheet,
    save to output_path.
    """
    # Open WITHOUT data_only to preserve Tracker formulas
    wb = openpyxl.load_workbook(str(input_path))
    sheet_name = cfg.excel.question_sheet_name
    if sheet_name not in wb.sheetnames:
        available_sheets = [s for s in wb.sheetnames if s not in cfg.excel.preserve_sheets]
        if available_sheets:
            sheet_name = available_sheets[0]
        else:
            raise ValueError(f"No valid sheets found in workbook. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    # Locate column indices
    header_row_idx = cfg.excel.header_row
    col_no       = _find_col(ws, header_row_idx, cfg.excel.no_column, ["No", "no", "Number", "number"])
    col_label    = _find_col(ws, header_row_idx, cfg.excel.question_column, ["Label", "label"])
    col_type     = _find_col(ws, header_row_idx, cfg.excel.type_column, ["Type", "type"])
    col_response = _find_col(ws, header_row_idx, cfg.excel.response_column, ["Response", "response"])
    col_existing_comments = _find_col(ws, header_row_idx, cfg.excel.existing_comments_column, ["Comments", "comments"])
    col_generated_comments = _find_col(ws, header_row_idx, cfg.excel.generated_comments_column, ["Comments from CB", "comments from cb"])
    col_comparison = _find_col(ws, header_row_idx, cfg.excel.comparison_column, ["Comparison and output", "comparison and output"])

    # Audit columns: H I J K L M (after Comparison column)
    col_confidence   = col_comparison + 1
    col_top_score    = col_comparison + 2
    col_sources      = col_comparison + 3
    col_needs_review = col_comparison + 4
    col_review_notes = col_comparison + 5
    col_revision_cnt = col_comparison + 6
    col_resp_len     = col_comparison + 7

    # Write headers in header row
    _write_cell(ws, header_row_idx, col_generated_comments, "Comments from CB")
    _write_cell(ws, header_row_idx, col_comparison, "Comparison and output")
    _write_cell(ws, header_row_idx, col_confidence,   "Confidence")
    _write_cell(ws, header_row_idx, col_top_score,    "Top Retrieval Score")
    _write_cell(ws, header_row_idx, col_sources,      "Sources")
    _write_cell(ws, header_row_idx, col_needs_review, "Needs Review")
    _write_cell(ws, header_row_idx, col_review_notes, "Review Notes")
    _write_cell(ws, header_row_idx, col_revision_cnt, "Revision Count")
    _write_cell(ws, header_row_idx, col_resp_len, "Response Length (Chars)")

    # Style the header row
    max_col = ws.max_column or col_resp_len
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        if cell.value is not None:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.border = _THIN_BORDER

    # Column widths
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(col_generated_comments)].width      = 60
    ws.column_dimensions[get_column_letter(col_review_notes)].width  = 50
    ws.column_dimensions[get_column_letter(col_comparison)].width    = 15
    ws.column_dimensions[get_column_letter(col_confidence)].width    = 15
    ws.column_dimensions[get_column_letter(col_top_score)].width     = 15
    ws.column_dimensions[get_column_letter(col_needs_review)].width  = 15
    ws.column_dimensions[get_column_letter(col_revision_cnt)].width  = 15
    ws.column_dimensions[get_column_letter(col_resp_len)].width      = 15

    wrap_align = Alignment(wrap_text=True, vertical="top")

    for row_number, fields in results.items():
        # Response (E) — Single Choice only
        if fields.get("single_choice_value"):
            _write_cell(ws, row_number, col_response, fields["single_choice_value"])

        # Generated Comments (F)
        answer_text = fields.get("answer_text", "")
        comments_cb_cell = ws.cell(row=row_number, column=col_generated_comments)
        comments_cb_cell.value = answer_text
        comments_cb_cell.alignment = wrap_align

        # Comparison (G)
        comparison_val = fields.get("comparison", "")
        comparison_cell = ws.cell(row=row_number, column=col_comparison)
        comparison_cell.value = comparison_val
        comparison_cell.alignment = wrap_align

        # Audit columns
        confidence = fields.get("confidence", "low")
        _write_cell(ws, row_number, col_confidence,   confidence)
        _write_cell(ws, row_number, col_top_score,    round(fields.get("top_retrieval_score", 0.0), 4))
        _write_cell(ws, row_number, col_sources,      ", ".join(fields.get("sources", [])))
        needs_review_str = "YES" if fields.get("needs_review") else "no"
        _write_cell(ws, row_number, col_needs_review, needs_review_str)

        review_notes_cell = ws.cell(row=row_number, column=col_review_notes)
        review_notes_cell.value = fields.get("review_notes", "")
        review_notes_cell.alignment = wrap_align

        _write_cell(ws, row_number, col_revision_cnt, fields.get("revision_count", 0))
        _write_cell(ws, row_number, col_resp_len, len(answer_text))

        # Color fills
        conf_cell = ws.cell(row=row_number, column=col_confidence)
        if confidence == "high":
            conf_cell.fill = _FILL_GREEN
        elif confidence == "medium":
            conf_cell.fill = _FILL_YELLOW
        else:
            conf_cell.fill = _FILL_RED

        if needs_review_str == "YES":
            comments_cb_cell.fill = _FILL_RED

        # Apply borders to output cells
        for c_idx in [col_response, col_generated_comments, col_comparison, col_confidence, col_top_score, col_sources, col_needs_review, col_review_notes, col_revision_cnt, col_resp_len]:
            ws.cell(row=row_number, column=c_idx).border = _THIN_BORDER

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Output saved to %s", output_path)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _find_col(ws: openpyxl.worksheet.worksheet.Worksheet,
              header_row: int,
              default_letter: str,
              candidates: list[str]) -> int:
    """Locate the column index by scanning the header row for candidate names."""
    max_col = ws.max_column or 26
    for col_idx in range(1, max_col + 1):
        val = _cell_str(ws, header_row, col_idx)
        if val in candidates:
            return col_idx
    # Fall back to default letter
    return column_index_from_string(default_letter)


def _cell_str(ws, row: int, col: int) -> str:
    val = ws.cell(row=row, column=col).value
    if val is None:
        return ""
    return str(val).strip()


def _write_cell(ws, row: int, col: int, value) -> None:
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.alignment = Alignment(wrap_text=True, vertical="top")
